from django.shortcuts import render,redirect,get_object_or_404
from django.http import HttpResponse
from django.http import JsonResponse
from .forms import *
from .models import *
import os
from django.contrib.auth import authenticate
from django.core.mail import send_mail
from datetime import date, datetime

from decimal import Decimal
from django.utils import timezone
from django.views.decorators.http import require_POST
import json
from dateutil import parser
from django.db.models import Count
from .models import regmodel,AttendanceModel,LeaveModel,ManagerModel
from .forms import logform,LeaveForm
# Create your views here.

####### EMPLOYEE #########

# Employee signin
from django.http import HttpResponse

def signin(request):
    if request.method == 'POST':
        form = logform(request.POST)
        if form.is_valid():
            employee_id = form.cleaned_data['employeeid']
            print("Employee ID:", employee_id) 

            password = form.cleaned_data['password']
            password = password.lower()

            employees = regmodel.objects.filter(employeeid=employee_id, password=password)

            if employees.exists():
                employee = employees.first()
                if employee.status in ['Active', 'Inactive']:  # Check if status is active or inactive
                    request.session['id'] = employee_id
                    print("Session ID:", request.session['id'])
                    return redirect(page2)  
                else:
                    # Show message for access denied if status is not active or inactive
                    return render(request, 'signin.html', {'form': logform(), 'popup_message': 'ACCESS DENIED.......!'})
    
    return render(request, 'signin.html', {'form': logform()})


# For calculating the login, logoff time
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.utils import timezone
from dateutil import parser
import json
from .models import AttendanceModel, ExtraModel

@require_POST
def handle_frontend_data(request):
    try:
        data = json.loads(request.body)
        userid = request.session.get('id')
        employee = regmodel.objects.get(employeeid=userid)
        name = employee.name

        # Convert time strings to datetime objects
        logintime_str = data.get('logintime')
        logintime = parser.parse(logintime_str)
        logofftime_str = data.get('logofftime')
        logofftime = parser.parse(logofftime_str)

        # Retrieve other time data from the request
        totalbreakTimex = data.get('totalbreakTimex')
        totalmeetingTimex = data.get('totalmeetingTimex')
        totaldownTimex = data.get('totaldownTimex')
        totalworkx = data.get('totalworkx')

        # Calculate working hours
        working_hours = (logofftime - logintime).total_seconds() / 3600

        # Get current date
        current_date = timezone.now().date()

        # Check if a record for this user and date already exists in AttendanceModel
        existing_attendance_record = AttendanceModel.objects.filter(employeeid=userid, created=current_date).first()

        if existing_attendance_record:
            # Update existing record
            existing_attendance_record.logintime = logintime
            existing_attendance_record.logofftime = logofftime
            existing_attendance_record.totalbreakTimex = totalbreakTimex
            existing_attendance_record.totalmeetingTimex = totalmeetingTimex
            existing_attendance_record.totaldownTimex = totaldownTimex
            existing_attendance_record.totalworkx = totalworkx
            existing_attendance_record.save()
        else:
            # Create new record
            AttendanceModel.objects.create(
                name=name,
                employeeid=userid,
                logintime=logintime,
                logofftime=logofftime,
                totalbreakTimex=totalbreakTimex,
                totalmeetingTimex=totalmeetingTimex,
                totaldownTimex=totaldownTimex,
                totalworkx=totalworkx,
                created=current_date
            )

        # Create or update ExtraModel record
        extra_model_record, created = ExtraModel.objects.get_or_create(
            employeeid=userid,
            date=current_date
        )

        # Determine the status based on working hours
        if working_hours >= 7:
            extra_model_record.status = 'Present'
        elif 5 <= working_hours < 7:
            extra_model_record.status = 'Half day'
        else:
            extra_model_record.status = 'Absent'

        extra_model_record.name = name
        extra_model_record.save()

        return JsonResponse({'status': 'success'})
    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': 'Invalid JSON data'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})


# Side and top navbar page of employee
def base(request):
    return render(request, 'base.html')


# Dashboard of employee
def page2(request):
    id1=request.session['id']
    a=regmodel.objects.get(employeeid=id1)
    img=str(a.image).split('/')[-1]
    return render(request,'dash.html',{'a':a})


# Profile page of employee
def page3(request):
    id1 = request.session.get('id')
    
    if id1:
        try:
            # Retrieve the employee details
            a = regmodel.objects.get(employeeid=id1)
            img = str(a.image).split('/')[-1]

            # Retrieve all departments for the manager
            managers = ManagerModel.objects.filter(employeeid=id1)

            # Check if the employee is a manager
            is_manager = managers.exists()

            if is_manager:
                # Fetch employees from regmodel for each department
                department_ids = [manager.department for manager in managers]
                employees = []
                for department_id in department_ids:
                    employees.extend(regmodel.objects.filter(department=department_id))
            else:
                employees = None

            return render(request, 'profilex.html', {'a': a, 'emp': employees, 'img': img, 'is_manager': is_manager})

        except regmodel.DoesNotExist:
            # Handle case when employee does not exist in regmodel
            return render(request, 'error.html', {'error_message': 'Employee does not exist.'})

    else:
        # Handle case when employee ID is not found in the session
        return render(request, 'error.html', {'error_message': 'Employee ID not found in session.'})


# Employee profile edit
def editprofile(request):
    id1 = request.session['id']
    a = regmodel.objects.get(employeeid=id1)
    img = str(a.image).split('/')[-1]
    return render(request, 'editprofile.html', {'a':a, 'img':img})



# Calendar display of employee
from django.shortcuts import render
from .models import regmodel, ExtraModel

def page4(request):
    id1 = request.session['id']
    a = regmodel.objects.get(employeeid=id1)
    attendance_records = ExtraModel.objects.filter(employeeid=id1).order_by('date')

    return render(request, 'attendancex.html', {'a': a, 'attendance_records': attendance_records})


def fetch_data(request, year, month):
    id1 = request.session['id']
    records = ExtraModel.objects.filter(employeeid=id1, date__year=year, date__month=month).order_by('date').values()
    return JsonResponse(list(records), safe=False)

def holidaycalendar(request):
    return render(request, 'holidaycalendar.html')


# Leave application of employee
from django.db.models import Q

from datetime import timedelta

def page5(request):
    id1 = request.session['id']
    employee = regmodel.objects.get(employeeid=id1)

    if request.method == 'POST':
        form = LeaveForm(request.POST)
        if form.is_valid():
            leave_instance = form.save(commit=False)
            leave_instance.applieddate = timezone.now()

            # Check if the 'half_day' radio button is selected
            if leave_instance.leaveType == 'Half Day':
                # Set the days to 0.5 for half-day leave
                leave_instance.days = 0.5
            else:
                # Retrieve 'from_date' and 'to_date' from the form
                from_date = form.cleaned_data['fromdate']
                to_date = form.cleaned_data['todate']

                # Check if both 'from_date' and 'to_date' are not None
                if from_date and to_date:
                    # Check if any public holidays fall within the specified range
                    public_holidays = PublicholidaysModel.objects.filter(date__range=[from_date, to_date])

                    # Check if any week-offs fall within the specified range
                    week_offs = WeekoffModel.objects.filter(
                        employeeid=id1,
                        weekoff1__range=[from_date, to_date],
                        weekoff2__range=[from_date, to_date]
                    )

                    # Create a set to store all excluded dates
                    excluded_dates = set()

                    # Exclude public holidays
                    for holiday in public_holidays:
                        excluded_dates.add(holiday.date)

                    # Exclude week-offs
                    for week_off in week_offs:
                        excluded_dates.add(week_off.weekoff1)
                        # Assuming weekoff2 is also a date field
                        excluded_dates.add(week_off.weekoff2)

                    # Calculate the difference in days, excluding public holidays and week-offs
                    days_difference = (to_date - from_date).days + 1

                    for date in excluded_dates:
                        if from_date <= date <= to_date:
                            days_difference -= 1

                    # Set the calculated days to the leave instance
                    leave_instance.days = days_difference

            # Associate the leave with the employee
            leave_instance.employee = employee

            # Save the leave instance to the database
            leave_instance.save()

            return redirect(page2)

    else:
        form = LeaveForm()

    return render(request, 'leaveapplicationx.html', {'a': employee, 'form': form})


# Leave status display for employee
from django.shortcuts import render
from django.http import JsonResponse
from .models import LeaveModel

def leavestatus(request):
    try:
        # Fetch the employee ID from the session
        id1 = request.session.get('id')

        # Check if the employee ID exists in the session
        if id1 is not None:
            # Fetch the leave status for the specific employee based on their ID
            leave_statuses = LeaveModel.objects.filter(employeeid=id1)

            # Pass the leave statuses to the template for rendering
            return render(request, 'leavestatus.html', {'leave_statuses': leave_statuses})
        else:
            # If the employee ID is not in the session, handle accordingly
            return JsonResponse({'status': 'error', 'message': 'Employee ID not found in session'})

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})
    
    
# # Weekoff display page for employee
# def weekoffemployee(request):
#     try:
#         # Fetch the employee ID from the session
#         id1 = request.session.get('id')

#         # Check if the employee ID exists in the session
#         if id1 is not None:
#             weekoff = ExtraModel.objects.filter(employeeid=id1, status='Weekoff')

#             return render(request, 'weekoffemployee.html', {'weekoff': weekoff})
#         else:
#             # If the employee ID is not in the session, handle accordingly
#             return JsonResponse({'status': 'error', 'message': 'Employee ID not found in session'})

#     except Exception as e:
#         return JsonResponse({'status': 'error', 'message': str(e)})
    
def weekoffemployee(request):
    try:
        # Fetch all unique dates from the ExtraModel where status is 'Weekoff'
        unique_dates = ExtraModel.objects.filter(status='Weekoff').values_list('date', flat=True).distinct()

        # Create a list to store the data for each date
        weekoff_data_by_date = []

        # Iterate over unique dates
        for date in unique_dates:
            # Fetch records for the current date and status 'Weekoff'
            records_for_date = ExtraModel.objects.filter(date=date, status='Weekoff')

            # Extract the name and employeeid for each record and store in a list
            records_data = [{'name': record.name, 'employeeid': record.employeeid} for record in records_for_date]

            # Append the data for the current date to the list
            weekoff_data_by_date.append({'date': date, 'records': records_data})

        # Pass the data to the template for rendering
        return render(request, 'weekoffemployee.html', {'weekoff_data_by_date': weekoff_data_by_date})

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})
    
    
# Public holidays display for employee    
def publicholidaysemployee(request):
    try:
        # Fetch all records from the AttendanceModel
        a =  (PublicholidaysModel.objects.exclude(status='Sunday')
               .values('date','status')
            .annotate(date_count=Count('date'))
            )

        # Pass the records to the template for rendering
        return render(request, 'publicholidaysemployee.html', {'a': a})

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})


# Salary slip of employee
from django.shortcuts import render
from .models import SalarySlipModel

def page6(request):
    if request.method == 'POST':
        # Get employee ID from session
        id1 = request.session.get('id')
        
        # Get selected month and year from the form data
        selected_month_year = request.POST.get('month-year-picker')
        selected_date = datetime.strptime(selected_month_year, '%Y-%m')
        
        # Filter the SalarySlipModel objects based on employee ID and selected month/year
        salary_slips = SalarySlipModel.objects.filter(employeeid=id1, month=selected_date.strftime('%B'), year=selected_date.year)
        
        # Fetch employee details
        employee = regmodel.objects.get(employeeid=id1)
        
        return render(request, 'salaryx.html', {'a': employee, 'salary_slips': salary_slips})
    else :
        return render(request, 'salaryx.html') 
    
# Productivity of employee    
from django.http import JsonResponse

def productivityemployee(request):
    try:
        # Fetch the employee ID from the session
        id1 = request.session.get('id')

        # Check if the employee ID exists in the session
        if id1 is not None:
            # Fetch the productivity data for the specific employee based on their ID
            productivity_data = ProductivityModel.objects.filter(employeeid=id1)

            # Pass the productivity data to the template for rendering
            return render(request, 'productivityemployee.html', {'productivity_data': productivity_data})
        else:
            # If the employee ID is not in the session, handle accordingly
            return JsonResponse({'status': 'error', 'message': 'Employee ID not found in session'})

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})
    
    
# Employee signout
def signout(request):
    # Check if the user is already logged in
    if 'id' in request.session:
        # If logged in, remove the user's session data
        del request.session['id']

    # Redirect to the desired logout success or home page
    return redirect('/signin/') 



########## ADMIN ###########

#Admin login
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate
from .forms import adminloginform # Make sure you import your form

def adminlogin(request):
    error_message = None  # Initialize error message
    if request.method == 'POST':
        a = adminloginform(request.POST)
        if a.is_valid():
            username = a.cleaned_data['username']
            password = a.cleaned_data['password']
            user = authenticate(request, username=username, password=password)
            if user is not None:
                return redirect(admindashboard)  # Use 'admindashboard' if it is the name of your URL pattern
            else:
                error_message = 'Incorrect username or password.'  # Set error message if authentication fails
    else:
        a = adminloginform()

    return render(request, 'adminlogin.html', {'form': a, 'error_message': error_message})



def managerlogin(request):
    if request.method == "POST":
        a=managerloginform(request.POST)
        if a.is_valid():
            username=a.cleaned_data['username']
            password=a.cleaned_data['password']
            b=ManagerModel.objects.all()
            for i in b:
                if i.username == username  and i.password == password:
                    return redirect(managerdashboard)
            else:
                return redirect(managerlogin)
    return render(request,'managerlogin.html')






# Admin dashboard
def admindashboard(request):
    return render(request,'admindashboard.html')

def managerdashboard(request):
    return render(request,'managerdashboard.html')


# Employee details display for admin
from django.shortcuts import render
from django.http import JsonResponse

from django.shortcuts import render, get_object_or_404
from .models import regmodel

from django.shortcuts import render, get_object_or_404, redirect
from .models import regmodel
from django.db.models import F

from django.core.exceptions import ObjectDoesNotExist



def employeedetailsman(request):
    # Retrieve the current user's employee ID
    current_user_id = request.session.get('id')
    
    if current_user_id:
        try:
            # Retrieve the current user's details
            current_user = regmodel.objects.get(employeeid=current_user_id)

            # Retrieve departments managed by the current user
            managed_departments = ManagerModel.objects.filter(employeeid=current_user_id).values_list('department', flat=True)

            # Fetch users from regmodel for the managed departments
            a = regmodel.objects.filter(department__in=managed_departments)

            return render(request, 'employeeman.html', {'a': a})

        except regmodel.DoesNotExist:
            # Handle case when current user does not exist
            pass
    else:
        # Handle case when employee ID is not found in the session
        pass

    return render(request, 'employeeman.html', {'a': []})



from django.shortcuts import render, get_object_or_404, redirect
from .models import regmodel, DepartmentModel
from django.core.exceptions import ObjectDoesNotExist

def employeedetails(request):
    a = regmodel.objects.filter(status__in=['Active', 'Inactive', 'Permission Denied']).order_by('-id')
    uid = []
    nme = []
    emid = []
    dsg = []
    eml = []
    blg = []
    phno = []
    db = []
    jnd = []
    img = []
    accno = []
    bknm = []
    brc = []
    ifs = []
    sl = []
    lgt = []
    lgt_sat = []
    adr = []
    s = []
    dep = []
    shf = []
    shf_sat = []
    idp = []
    ed = []
    wrk = []
    re = []
    ot = []
    ms = []
    gn = []

    for i in a:
        id = i.id
        uid.append(id)
        nm = i.name
        nme.append(nm)
        eid = i.employeeid
        emid.append(eid)
        ds = i.designation
        dsg.append(ds)
        em = i.email
        eml.append(em)
        bg = i.bloodgroup
        blg.append(bg)
        phn = i.phonenumber
        phno.append(phn)
        dob = i.dateofbirth
        db.append(dob)
        jn = i.joiningdate
        jnd.append(jn)
        im = str(i.image).split('/')[-1]
        img.append(im)
        acno = i.accountnumber
        accno.append(acno)
        bnknm = i.bankname
        bknm.append(bnknm)
        br = i.branch
        brc.append(br)
        ifsc = i.ifsccode
        ifs.append(ifsc)
        sal = i.salary
        sl.append(sal)
        lg = i.logintime
        lgt.append(lg)
        lgtm = i.logintime_sat
        lgt_sat.append(lgtm)
        add = i.address
        adr.append(add)
        sts = i.status
        s.append(sts)

        # Fetch department name from DepartmentModel
        department_id = i.department
        try:
            department = DepartmentModel.objects.get(id=department_id)
            de = department.department
        except ObjectDoesNotExist:
            de = ""
        dep.append(de)

        sh = i.shifttime
        shf.append(sh)
        sht = i.shifttime_sat
        shf_sat.append(sht)
        idpr = i.idproof
        idp.append(idpr)
        edu = i.educationalcertificate
        ed.append(edu)
        wrke = i.workexperience
        wrk.append(wrke)
        res = i.resume
        re.append(res)
        oth = i.others
        ot.append(oth)

        mrs = i.maritalstatus
        ms.append(mrs)
        gen = i.gender
        gn.append(gen)

    pair = zip(uid, nme, emid, dsg, eml, blg, phno, db, jnd, img, accno, bknm, brc, ifs, sl, lgt, lgt_sat, adr, s, dep, shf, shf_sat, idp, ed, wrk, re, ot, ms, gn)

    if request.method == 'POST':
        employee_id = request.POST.get('employee_id')
        action = request.POST.get('action')

        employee = get_object_or_404(regmodel, id=employee_id)

        if action == 'inactive':
            employee.status = 'Active' if employee.status == 'Inactive' else 'Inactive'
        elif action == 'active':
            employee.status = 'Inactive' if employee.status == 'Active' else 'Active'
        elif action == 'resigned':
            employee.status = 'Resigned'
        elif action == 'terminated':
            employee.status = 'Terminated'
        elif action == 'deny_permission':
            employee.status = 'Permission Denied'
        elif action == 'grant_permission':
            employee.status = 'Active'
        else:
            # Handle invalid action, if needed
            pass
        
        # Save the updated status
        employee.save()
        
        return redirect(employeedetails)

    return render(request, 'employeedetails.html', {'a': pair})


from django.http import JsonResponse
from .models import DepartmentModel

def save_department(request):
    if request.method == 'POST':
        department_name = request.POST.get('department_name')
        department = DepartmentModel.objects.create(department=department_name)
        return JsonResponse({'success': True})  # Return JSON response indicating success
    else:
        return JsonResponse({'success': False, 'error': 'Invalid request method'})




from django.http import JsonResponse
from .models import DepartmentModel

def get_departments(request):
    departments = DepartmentModel.objects.all()
    data = [{'name': department.department} for department in departments]
    return JsonResponse({'departments': data})



# Employee registration for admin
from django.shortcuts import render, redirect
from django.http import HttpResponse
from .models import ExtraModel, regmodel
from datetime import timedelta
from django.db.models import Q

from django.shortcuts import render, redirect
from django.utils import timezone
from django.db.models import Q
from .models import PublicholidaysModel, ExtraModel, regmodel
from .forms import regform

from django.shortcuts import render, redirect
from django.http import HttpResponse
from .forms import regform
from .models import regmodel, ExtraModel
from django.db.models import Q

from django.shortcuts import render, redirect, HttpResponse
from django.db.models import Q
from .models import regmodel
from .forms import regform 
from .models import ExtraModel, DepartmentModel

def register(request):
    if request.method == 'POST':
        form = regform(request.POST, request.FILES)
        if form.is_valid():
            # Extracting form data
            nm = form.cleaned_data['name']
            eid = form.cleaned_data['employeeid']
            ds = form.cleaned_data['designation']
            em = form.cleaned_data['email']
            bg = form.cleaned_data['bloodgroup']
            phn = form.cleaned_data['phonenumber']
            dob = form.cleaned_data['dateofbirth']
            jnd = form.cleaned_data['joiningdate']
            img = form.cleaned_data['image']
            acno = form.cleaned_data['accountnumber']
            bnknm = form.cleaned_data['bankname']
            br = form.cleaned_data['branch']
            ifsc = form.cleaned_data['ifsccode']
            sal = form.cleaned_data['salary']
            psw = form.cleaned_data['password']
            cpsw = form.cleaned_data['confirmpassword']
            lgt = form.cleaned_data['logintime']  
            lgt_sat = form.cleaned_data.get('logintime_sat', None)  
            add = form.cleaned_data['address']
            dep = form.cleaned_data['department']
            shf = form.cleaned_data['shifttime']
            shf_sat = form.cleaned_data.get('shifttime_sat')
            id = form.cleaned_data['idproof']
            ed = form.cleaned_data['educationalcertificate']
            wrk = form.cleaned_data['workexperience']
            res = form.cleaned_data['resume']
            oth = form.cleaned_data['others']
            ms = form.cleaned_data['maritalstatus']
            gn = form.cleaned_data['gender']

            if psw == cpsw:
                # Saving employee data
                employee = regmodel(
                    name=nm, employeeid=eid, designation=ds, email=em, bloodgroup=bg,
                    phonenumber=phn, dateofbirth=dob, joiningdate=jnd, image=img, accountnumber=acno,
                    bankname=bnknm, branch=br, ifsccode=ifsc, salary=sal, password=psw, logintime=lgt, 
                    logintime_sat=lgt_sat, 
                    address=add, department=dep, shifttime=shf, shifttime_sat=shf_sat, idproof=id, educationalcertificate=ed,
                    workexperience=wrk, resume=res, others=oth, maritalstatus=ms, gender=gn
                )
                employee.save()

                # Fetch existing holidays and Sundays from ExtraModel
                existing_dates = ExtraModel.objects.filter(Q(status='Holiday') | Q(status='Sunday')).values_list('date', 'status')

                # If there are existing dates, copy them for the newly registered employee
                if existing_dates.exists():
                    for date, status in existing_dates:
                        # Check if the date already exists for this employee, to avoid duplication
                        if not ExtraModel.objects.filter(employeeid=eid, date=date).exists():
                            # Save the existing date for the newly registered employee with the correct status
                            extra_data = ExtraModel(name=nm, employeeid=eid, date=date, status=status)
                            extra_data.save()

                return redirect(employeedetails)  # Replace with your actual URL name
            else:
                return HttpResponse("Password doesn't match")
        else:
            return HttpResponse("Registration failed.")
    else:
        form = regform()

    # Fetching list of employees for reporting manager dropdown
    employees = regmodel.objects.filter(status__in=['Active', 'Inactive'])
    departments = DepartmentModel.objects.all()
    return render(request, 'register.html', {'form': form, 'employees': employees, 'departments': departments})



#  Resigned employees display for admin
from django.shortcuts import render
from .models import regmodel

def resigned_employees(request):
    resigned_employees = regmodel.objects.filter(status='Resigned')
    return render(request, 'resigned.html', {'resigned_employees': resigned_employees})


def resigned_employeesmanager(request):
    resigned_employees = regmodel.objects.filter(status='Resigned')
    return render(request, 'resignedmanager.html', {'resigned_employees': resigned_employees})

# Calendar display for admin
from decimal import Decimal
from django.shortcuts import render, HttpResponse, get_object_or_404
from .models import ExtraModel, regmodel, SalarySlipModel

from decimal import Decimal, InvalidOperation

from django.http import JsonResponse

from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, get_object_or_404
from .models import ExtraModel, SalarySlipModel

def view_monthly_details(request, employee_id):
    request.session['empid'] = employee_id
    employees = ExtraModel.objects.filter(employeeid=employee_id).values('name', 'employeeid').distinct()

    if employees.exists():
        employee_salary = get_object_or_404(regmodel, employeeid=employee_id).salary

        if request.method == 'POST':
            name = request.POST.get('name', '')
            month = request.POST.get('month', '')
            year = request.POST.get('year', '')

            total_working_days = convert_to_decimal(request.POST.get('totalPayableDays', 0))
            total_leave = convert_to_decimal(request.POST.get('totalleave', 0))
            paid_leave = convert_to_decimal(request.POST.get('paidleave', 0))
            late_login = convert_to_decimal(request.POST.get('Latelogin', 0))
            salary_deducted_leave = convert_to_decimal(request.POST.get('salarydeductedleave', 0))
            salary = convert_to_decimal(request.POST.get('salary', 0))
            per_day_salary = convert_to_decimal(request.POST.get('perdaysalary', 0))
            deduction_amount = convert_to_decimal(request.POST.get('deductionamount', 0))
            monthly_salary = convert_to_decimal(request.POST.get('monthlysalary', 0))

            incentive = convert_to_decimal(request.POST.get('incentive', 0))
            leave_encashment = convert_to_decimal(request.POST.get('leaveencashment', 0))
            byod = convert_to_decimal(request.POST.get('byod', 0))

            existing_entry = SalarySlipModel.objects.filter(
                employeeid=employee_id,
                month=month,
                year=year
            ).first()

            if existing_entry:
                existing_entry.name = name
                existing_entry.totalPayableDays = total_working_days
                existing_entry.totalleave = total_leave
                existing_entry.paidleave = paid_leave
                existing_entry.Latelogin = late_login
                existing_entry.salarydeductedleave = salary_deducted_leave
                existing_entry.salary = salary
                existing_entry.perdaysalary = per_day_salary
                existing_entry.deductionamount = deduction_amount
                existing_entry.incentive = incentive
                existing_entry.leaveencashment = leave_encashment
                existing_entry.byod = byod  # Update BYOD
                existing_entry.monthlysalary = monthly_salary
                existing_entry.save()
            else:
                salary_slip = SalarySlipModel(
                    name=name,
                    employeeid=employee_id,
                    month=month,
                    year=year,
                    totalPayableDays=total_working_days,
                    totalleave=total_leave,
                    paidleave=paid_leave,
                    Latelogin=late_login,
                    salarydeductedleave=salary_deducted_leave,
                    salary=salary,
                    perdaysalary=per_day_salary,
                    deductionamount=deduction_amount,
                    incentive=incentive,
                    leaveencashment=leave_encashment,
                    byod=byod,  # Add BYOD
                    monthlysalary=monthly_salary
                )
                salary_slip.save()

            return JsonResponse({'message': "Salary Slip generated and saved successfully!"})

        return render(request, 'calendar.html', {'employees': employees, 'salary': employee_salary})
    else:
        return HttpResponse("Details do not exist!!!")



def convert_to_int(value):
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return 0

def convert_to_decimal(value):
    try:
        return Decimal(value)
    except (InvalidOperation, ValueError):
        return Decimal(0)

def calendarman(request, employee_id):
    request.session['empid'] = employee_id
    employees = ExtraModel.objects.filter(employeeid=employee_id).values('name', 'employeeid').distinct()

    if employees.exists():
        employee_salary = get_object_or_404(regmodel, employeeid=employee_id).salary

        if request.method == 'POST':
            name = request.POST.get('name', '')
            month = request.POST.get('month', '')
            year = request.POST.get('year', '')

            total_working_days = convert_to_decimal(request.POST.get('totalPayableDays', 0))
            total_leave = convert_to_decimal(request.POST.get('totalleave', 0))
            paid_leave = convert_to_decimal(request.POST.get('paidleave', 0))
            late_login = convert_to_decimal(request.POST.get('Latelogin', 0))
            salary_deducted_leave = convert_to_decimal(request.POST.get('salarydeductedleave', 0))
            salary = convert_to_decimal(request.POST.get('salary', 0))
            per_day_salary = convert_to_decimal(request.POST.get('perdaysalary', 0))
            deduction_amount = convert_to_decimal(request.POST.get('deductionamount', 0))
            monthly_salary = convert_to_decimal(request.POST.get('monthlysalary', 0))

            # Check if incentive is provided
            incentive = Decimal(0)
            if 'incentive' in request.POST:
                incentive = convert_to_decimal(request.POST.get('incentive', 0))

            # Check if leave encashment is provided
            leave_encashment = Decimal(0)
            if 'leaveencashment' in request.POST:
                leave_encashment = convert_to_decimal(request.POST.get('leaveencashment', 0))

            # Check if data already exists for the given employee and month
            existing_entry = SalarySlipModel.objects.filter(
                employeeid=employee_id,
                month=month,
                year=year
            ).first()

            if existing_entry:
                # Update existing entry
                existing_entry.name = name
                existing_entry.totalPayableDays = total_working_days
                existing_entry.totalleave = total_leave
                existing_entry.paidleave = paid_leave
                existing_entry.Latelogin = late_login
                existing_entry.salarydeductedleave = salary_deducted_leave
                existing_entry.salary = salary
                existing_entry.perdaysalary = per_day_salary
                existing_entry.deductionamount = deduction_amount
                existing_entry.incentive = incentive
                existing_entry.leaveencashment = leave_encashment  # Update leave encashment
                existing_entry.monthlysalary = monthly_salary
                existing_entry.save()
            else:
                # Insert new entry
                salary_slip = SalarySlipModel(
                    name=name,
                    employeeid=employee_id,
                    month=month,
                    year=year,
                    totalPayableDays=total_working_days,
                    totalleave=total_leave,
                    paidleave=paid_leave,
                    Latelogin=late_login,
                    salarydeductedleave=salary_deducted_leave,
                    salary=salary,
                    perdaysalary=per_day_salary,
                    deductionamount=deduction_amount,
                    incentive=incentive,
                    leaveencashment=leave_encashment,  # Add leave encashment
                    monthlysalary=monthly_salary
                )
                salary_slip.save()

            return JsonResponse({'message': "Salary Slip generated and saved successfully!"})

        return render(request, 'calendarman.html', {'employees': employees, 'salary': employee_salary})
    else:
        return HttpResponse("Details do not exist!!!")

def fetch_dat(request, year, month):
    empid = request.session.get('empid', None)  
    print("employee id =", empid)

    # Calculate start and end dates for the selected month
    start_date = date(year, month, 26)
    if month == 12:
        end_date = date(year , month , 25)
    else:
        end_date = date(year, month, 25)

    # Calculate start date of the previous month
    if month == 1:
        prev_month_start = date(year - 1, 12, 26)
    else:
        prev_month_start = date(year, month - 1, 26)

    # Fetch data for the selected month
    records_current_month = ExtraModel.objects.filter(employeeid=empid, date__range=(prev_month_start, end_date)).order_by('date').values()

    # Fetch data for the specified year and month
    records_specified_month = ExtraModel.objects.filter(employeeid=empid, date__year=year, date__month=month).order_by('date').values()
    
   # records = ExtraModel.objects.filter(employeeid=empid, date__year=year, date__month=month).order_by('date').values()
    # Combine the results
    records_current_month_list = list(records_current_month)
    records_specified_month_list = list(records_specified_month)

    # Create a dictionary to hold both lists
    response_data = {
        'records_current_month': records_current_month_list,
        'records_specified_month': records_specified_month_list,
    }

    return JsonResponse(response_data, safe=False)


# Employee attendance edit by admin
from django.shortcuts import render, redirect, get_object_or_404
from .models import ExtraModel
from .forms import ExtraForm

from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse

from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from .models import ExtraModel

from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from .models import ExtraModel

def editattendance(request, employee_id):
    extra_instance = get_object_or_404(ExtraModel, id=employee_id)

    if request.method == 'POST':
        extra_instance.status = request.POST.get('status')
        extra_instance.login = request.POST.get('login')
        extra_instance.comments = request.POST.get('comments', '')  # Default to an empty string if not provided

        # Adjust the login field if necessary
        if extra_instance.login == 'Latelogin' and extra_instance.status == 'Ontime':
            extra_instance.login = 'Ontime'

        extra_instance.save()

        employeeid = request.POST.get('employeeid')
        return redirect(reverse('view_monthly_details', args=[employeeid]))

    return render(request, 'editattendance.html', {'employee_details': extra_instance})










import os
from django.shortcuts import render, redirect
from .models import regmodel

from django.shortcuts import render, redirect
import os
from .models import regmodel, ExtraModel, DepartmentModel

from django.shortcuts import render, redirect, get_object_or_404
import os
from .models import regmodel, ExtraModel, DepartmentModel  # Adjust import according to your project structure

def editemployeedetails(request, id):
    a = get_object_or_404(regmodel, id=id)
    img = str(a.image).split('/')[-1]

    if request.method == 'POST':
        old_name = a.name  # Store the old name to find in ExtraModel
        old_employeeid = a.employeeid  # Store the old employeeid to find in ExtraModel

        a.name = request.POST.get('name')
        a.employeeid = request.POST.get('employeeid')
        a.designation = request.POST.get('designation')
        a.email = request.POST.get('email')
        a.bloodgroup = request.POST.get('bloodgroup')
        a.phonenumber = request.POST.get('phonenumber')
        a.dateofbirth = request.POST.get('dateofbirth')
        a.joiningdate = request.POST.get('joiningdate')
        a.accountnumber = request.POST.get('accountnumber')
        a.bankname = request.POST.get('bankname')
        a.branch = request.POST.get('branch')
        a.ifsccode = request.POST.get('ifsccode')
        a.salary = request.POST.get('salary')
        a.logintime = request.POST.get('logintime')
        a.logintime_sat = request.POST.get('logintime_sat')
        a.address = request.POST.get('address')
        a.department = request.POST.get('department')
        a.shifttime = request.POST.get('shifttime')
        a.shifttime_sat = request.POST.get('shifttime_sat')
        a.maritalstatus = request.POST.get('maritalstatus')
        a.gender = request.POST.get('gender')

        # Handling file uploads
        if 'image' in request.FILES:
            if a.image:
                os.remove(a.image.path)
            a.image = request.FILES['image']

        # Update PDF files
        for field_name in ['idproof', 'educationalcertificate', 'workexperience', 'resume', 'others']:
            if field_name in request.FILES:
                file_field = request.FILES[field_name]
                if getattr(a, field_name):
                    os.remove(getattr(a, field_name).path)
                setattr(a, field_name, file_field)

        a.save()

        # Update the name and employeeid in ExtraModel
        extra_entries = ExtraModel.objects.filter(employeeid=old_employeeid)
        for entry in extra_entries:
            entry.name = a.name
            entry.employeeid = a.employeeid
            entry.save()

        return redirect(employeedetails)
    
    # Fetching list of employees for reporting manager dropdown
    employees = regmodel.objects.filter(status__in=['Active', 'Inactive'])
    departments = DepartmentModel.objects.all()
    return render(request, 'editemployeedetails.html', {'a': a, 'img': img, 'employees': employees, 'departments': departments})





# Productivity entering for employee by admin
from django.shortcuts import render, redirect, get_object_or_404
from .models import ProductivityModel

def productivity(request, employee_id):
    if request.method == 'POST':
        # Get data from the form
        name = request.POST.get('name')
        employeeid = request.POST.get('employeeid')
        month = request.POST.get('month')
        productivity = request.POST.get('productivity')
        quality = request.POST.get('quality')
        appreciations = request.POST.get('appreciations')
        extra_initiatives = request.POST.get('extraInitiatives')
        target = request.POST.get('target')
        achievement = request.POST.get('achievement')
        percentage = request.POST.get('percentage')
        new_client = request.POST.get('newClient')
        renewals = request.POST.get('renewals')

        # Check if there is already an entry for this employeeid and month
        existing_entry = ProductivityModel.objects.filter(employeeid=employeeid, month=month).first()

        if existing_entry:
            # Update existing entry
            existing_entry.name = name
            existing_entry.productivity = productivity
            existing_entry.quality = quality
            existing_entry.appreciations = appreciations
            existing_entry.extraInitiatives = extra_initiatives
            existing_entry.target = target
            existing_entry.achievement = achievement
            existing_entry.percentage = percentage
            existing_entry.newClient = new_client
            existing_entry.renewals = renewals
            existing_entry.save()
        else:
            # Create new entry
            productivity_instance = ProductivityModel(
                name=name,
                employeeid=employeeid,
                month=month,
                productivity=productivity,
                quality=quality,
                appreciations=appreciations,
                extraInitiatives=extra_initiatives,
                target=target,
                achievement=achievement,
                percentage=percentage,
                newClient=new_client,
                renewals=renewals
            )
            productivity_instance.save()

        # Redirect after successful submission
        return redirect(admindashboard)  # Replace 'admindashboard' with your actual dashboard URL name

    # Render the form template for GET requests
    return render(request, 'productivity.html', {'employee_id': employee_id})



def productivityman(request, employee_id):
    if request.method == 'POST':
        print(request.POST) 
        # Get data from the form
        name = request.POST.get('name')
        employeeid = request.POST.get('employeeid')
        month = request.POST.get('month')
        productivity = request.POST.get('productivity')
        quality = request.POST.get('quality')
        appreciations = request.POST.get('appreciations')
        extra_initiatives = request.POST.get('extraInitiatives')
        target = request.POST.get('target')
        achievement = request.POST.get('achievement')
        percentage = request.POST.get('percentage')
        new_client = request.POST.get('newClient')
        renewals = request.POST.get('renewals')

        # Create and save ProductivityModel instance
        productivity_instance = ProductivityModel(
            name=name,
            employeeid=employeeid,
            month=month,
            productivity=productivity,
            quality=quality,
            appreciations=appreciations,
            extraInitiatives=extra_initiatives,
            target=target,
            achievement=achievement,
            percentage=percentage,
            newClient=new_client,
            renewals=renewals
        )
        productivity_instance.save()

        # Redirect after successful submission
        return redirect(page2)

    # Render the form template for GET requests
    return render(request, 'productivityman.html', {'employee_id': employee_id})


# Excel data upload by admin
from django.shortcuts import render, redirect
from django.http import HttpResponse
from .forms import ExcelUploadForm
from .models import ExcelModel
from datetime import datetime
from openpyxl import load_workbook

from django.http import HttpResponse
from openpyxl import load_workbook
from .models import ExcelModel, ExtraModel, regmodel
from datetime import datetime, date, time

def display_attendance(request):
    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']

            # Load the Excel workbook using openpyxl
            workbook = load_workbook(excel_file)

            # Assuming the data is in the first sheet
            worksheet = workbook.active

            # Initialize lists to store problematic rows
            invalid_rows = []

            # Iterate over rows in the worksheet, skipping the header row
            for idx, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
                if idx == 1:  # Skip the header row
                    continue

                # Check if the row is entirely None
                if all(cell is None for cell in row):
                    continue

                try:
                    # Assuming the first element is 'No' and not needed
                    _, employeeid, name, date, intime, outtime = row

                    # Ensure none of the essential fields are None
                    if None in (employeeid, name, date, intime, outtime):
                        raise ValueError("One of the essential fields is None")
                    
                    # Check if the date is already a string
                    if not isinstance(date, str):
                        # Convert the datetime.datetime object to a string
                        date = date.strftime('%d %B %Y')
                    
                    # Parse the date string into YYYY-MM-DD format
                    parsed_date = datetime.strptime(date, '%d %B %Y').strftime('%Y-%m-%d')
                    parsed_date_obj = datetime.strptime(parsed_date, '%Y-%m-%d').date()

                    # Create an instance of ExcelModel and save it to the database
                    ExcelModel.objects.create(
                        employeeid=employeeid,
                        name=name,
                        date=parsed_date,
                        intime=intime,
                        outtime=outtime
                    )
                except ValueError:
                    # Add the row to the list of invalid rows
                    invalid_rows.append(row)

            # Check if there are any invalid rows
            if invalid_rows:
                error_message = f"Some rows contain invalid data: {invalid_rows}"
                return HttpResponse(error_message, status=400)

            return HttpResponse('Data uploaded successfully!!!')
    else:
        form = ExcelUploadForm()

    return render(request, 'attendance_template.html', {'form': form})




# Excel data display by admin
from .models import ExcelModel, ExtraModel, regmodel
from django.shortcuts import render
from datetime import datetime, date, time, timedelta

from .models import ExcelModel, ExtraModel, regmodel
from django.shortcuts import render
from datetime import datetime, date, time, timedelta

from .models import ExcelModel, ExtraModel, regmodel
from django.shortcuts import render
from datetime import datetime, date, time, timedelta

def calculate_working_hours(intime, outtime):
    if intime and outtime:
        delta = datetime.combine(date.today(), outtime) - datetime.combine(date.today(), intime)
        hours, remainder = divmod(delta.seconds, 3600)
        minutes = remainder // 60
        working_hours = f"{hours} hours {minutes} minutes"
        return working_hours
    else:
        return "N/A"

def get_attendance_status(working_hours, is_saturday):
    if working_hours != "N/A":
        # Split the working_hours string into hours and minutes
        hours_str, minutes_str = working_hours.split(' hours ')
        
        # Convert hours and minutes to integers
        hours = int(hours_str)
        minutes = int(minutes_str.split(' minutes')[0])

        # Calculate total minutes
        total_minutes = hours * 60 + minutes

        if is_saturday:
            if total_minutes >= 420:
                return "Present"
            elif 240 <= total_minutes < 420:
                return "Half Day Leave"
            else:
                return "Leave"
        else:
            if total_minutes >= 480:
                return "Present"
            elif 300 <= total_minutes < 480:
                return "Half Day Leave"
            else:
                return "Leave"
    else:
        return "N/A"

from datetime import datetime

def display_attendance_details(request):
    # Get all data ordered by date
    excel_data = ExcelModel.objects.all().order_by('-date')

    # Iterate over the data and process as needed
    for row in excel_data:
        row.working_hours = calculate_working_hours(row.intime, row.outtime)
        
        # Check if the date is a Saturday
        is_saturday = row.date.weekday() == 5  # 5 represents Saturday
        
        if is_saturday:
            # Fetch regular login time for Saturday (logintime_sat)
            reg_data = regmodel.objects.filter(employeeid=row.employeeid).values('logintime_sat').first()
            if reg_data and reg_data['logintime_sat']:
                logintime_sat_str = reg_data['logintime_sat']
                logintime_sat = datetime.strptime(logintime_sat_str, '%H:%M').time()
                row.login = 'Latelogin' if row.intime > logintime_sat else ''
            else:
                row.login = ''
        else:
            # Fetch regular login time (logintime)
            reg_data = regmodel.objects.filter(employeeid=row.employeeid).values('logintime').first()
            if reg_data and reg_data['logintime']:
                logintime_str = reg_data['logintime']
                logintime = datetime.strptime(logintime_str, '%H:%M').time()
                row.login = 'Latelogin' if row.intime > logintime else ''
            else:
                row.login = ''
        
        row.status = get_attendance_status(row.working_hours, is_saturday)
        
        existing_data = ExtraModel.objects.filter(employeeid=row.employeeid, date=row.date).first()
        
        if not existing_data:  # Only add new records if they don't exist
            reg_data = regmodel.objects.filter(employeeid=row.employeeid).values('name').first()
            
            extra_data = ExtraModel(
                name=reg_data['name'] if reg_data and reg_data['name'] else row.name,
                employeeid=row.employeeid,
                date=row.date,
                status=row.status,
                login=row.login
            )
            extra_data.save()

    return render(request, 'attendancedetails.html', {'excel_data': excel_data})






# Leave display for admin
def leavedisplay(request):
    a = LeaveModel.objects.filter(status='Pending')
    uid = []
    nm = []
    emp = []
    apd = []

    for i in a:
        id = i.id
        uid.append(id) 
        nme = i.name
        nm.append(nme)
        em = i.employeeid
        emp.append(em)
        ap = i.applieddate
        apd.append(ap)

    pair = zip(uid, nm, emp, apd)
    return render(request, 'leavedisplay.html', {'a': pair})

from django.shortcuts import render
from .models import LeaveModel, ManagerModel, regmodel

def leavemanager(request):
    current_user_id = request.session.get('id')
    if current_user_id:
        try:
            current_user = ManagerModel.objects.get(employeeid=current_user_id)
            managed_department = current_user.department
            
            # Fetch employees of the managed department
            managed_employees = regmodel.objects.filter(department=managed_department)
            
            # Filter pending leaves for managed employees
            pending_leaves = LeaveModel.objects.filter(employeeid__in=managed_employees.values_list('employeeid', flat=True), status='Pending')

            return render(request, 'leavemanager.html', {'a': pending_leaves})

        except ManagerModel.DoesNotExist:
            pass
    else:
        pass

    return render(request, 'leavemanager.html', {'a': []})




# Leave details display and status update for admin
from django.shortcuts import render, redirect
from datetime import timedelta
from .models import LeaveModel, ExtraModel

def leave(request, id):
    a = LeaveModel.objects.get(id=id)

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'approve':
            a.status = 'Approved'
            
            # Update other fields if needed
            a.name = request.POST.get('name', a.name)
            a.employeeid = request.POST.get('employeeid', a.employeeid)
            a.fromdate = request.POST.get('fromdate', a.fromdate)
            a.todate = request.POST.get('todate', a.todate)
            a.reason = request.POST.get('reason', a.reason)
            a.days = request.POST.get('days', a.days)

            # Save the updated leave status and other fields
            a.save()

            # If the action is 'approve' and it is a half-day leave, create a new ExtraModel instance
            if a.days == 0.5:
                extra_instance = ExtraModel(
                    name=a.name,
                    employeeid=a.employeeid,
                    date=a.halfDayDate,
                    status='Half Day Leave'
                )
                extra_instance.save()
            else:
                # If it's a full-day leave, create a new ExtraModel instance for each date in the range
                from_date = a.fromdate
                to_date = a.todate

                while from_date <= to_date:
                    # Check if the leave date already exists in ExtraModel
                    existing_entry = ExtraModel.objects.filter(
                        employeeid=a.employeeid,
                        date=from_date
                    ).exists()

                    if not existing_entry:
                        # If the date doesn't exist, create a new ExtraModel instance
                        extra_instance = ExtraModel(
                            name=a.name,
                            employeeid=a.employeeid,
                            date=from_date,
                            status='Leave'
                        )
                        extra_instance.save()

                    # Move to the next date
                    from_date += timedelta(days=1)

            # Redirect to the admin dashboard after processing the leave request
            return redirect(admindashboard)  

        elif action == 'reject':
            a.status = 'Rejected'
            
            # Update other fields if needed
            a.name = request.POST.get('name', a.name)
            a.employeeid = request.POST.get('employeeid', a.employeeid)
            a.fromdate = request.POST.get('fromdate', a.fromdate)
            a.todate = request.POST.get('todate', a.todate)
            a.reason = request.POST.get('reason', a.reason)
            a.days = request.POST.get('days', a.days)

            # Save the updated leave status and other fields
            a.save()

            # Redirect to the admin dashboard after processing the leave request
            return redirect(admindashboard)  

    return render(request, 'leave.html', {'a': a})


def leaveman(request, id):
    a = LeaveModel.objects.get(id=id)

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'approve':
            a.status = 'Approved'
            
            # Update other fields if needed
            a.name = request.POST.get('name', a.name)
            a.employeeid = request.POST.get('employeeid', a.employeeid)
            a.fromdate = request.POST.get('fromdate', a.fromdate)
            a.todate = request.POST.get('todate', a.todate)
            a.reason = request.POST.get('reason', a.reason)
            a.days = request.POST.get('days', a.days)

            # Save the updated leave status and other fields
            a.save()

            # If the action is 'approve' and it is a half-day leave, create a new ExtraModel instance
            if a.days == 0.5:
                extra_instance = ExtraModel(
                    name=a.name,
                    employeeid=a.employeeid,
                    date=a.halfDayDate,
                    status='Half Day Leave'
                )
                extra_instance.save()
            else:
                # If it's a full-day leave, create a new ExtraModel instance for each date in the range
                from_date = a.fromdate
                to_date = a.todate

                while from_date <= to_date:
                    # Check if the leave date already exists in ExtraModel
                    existing_entry = ExtraModel.objects.filter(
                        employeeid=a.employeeid,
                        date=from_date
                    ).exists()

                    if not existing_entry:
                        # If the date doesn't exist, create a new ExtraModel instance
                        extra_instance = ExtraModel(
                            name=a.name,
                            employeeid=a.employeeid,
                            date=from_date,
                            status='Leave'
                        )
                        extra_instance.save()

                    # Move to the next date
                    from_date += timedelta(days=1)

            # Redirect to the admin dashboard after processing the leave request
            return redirect(page3)  

        elif action == 'reject':
            a.status = 'Rejected'
            
            # Update other fields if needed
            a.name = request.POST.get('name', a.name)
            a.employeeid = request.POST.get('employeeid', a.employeeid)
            a.fromdate = request.POST.get('fromdate', a.fromdate)
            a.todate = request.POST.get('todate', a.todate)
            a.reason = request.POST.get('reason', a.reason)
            a.days = request.POST.get('days', a.days)

            # Save the updated leave status and other fields
            a.save()

            # Redirect to the admin dashboard after processing the leave request
            return redirect(page3)  

    return render(request, 'leaveman.html', {'a': a})


# Applied leaves display for admin   
def appliedleaves(request):
    try:
        # Fetch all records from the AttendanceModel
        appliedleaves = LeaveModel.objects.all()

        # Pass the records to the template for rendering
        return render(request, 'appliedleaves.html', {'appliedleaves': appliedleaves})

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})
    
# Weekoff enetering for employee by admin    
from django.shortcuts import render, redirect
from .models import WeekoffModel, ExtraModel
from .forms import WeekoffForm
from django.utils import timezone
from django.shortcuts import get_object_or_404
from django.shortcuts import render, redirect
from .models import ExtraModel
from .forms import ExtraForm  



def weekoff(request):
    if request.method == 'POST':
        date = request.POST.get('date') 

        # Get the list of selected employee IDs from the checkboxes
        selected_employee_ids = request.POST.getlist('employee_checkbox')

        for employee_id in selected_employee_ids:
            form_data = {
                'name': request.POST.get(f'name_{employee_id}'),
                'employeeid': request.POST.get(f'employeeid_{employee_id}'),
                'date': date,  
                'status': 'Weekoff',
            }

            form = ExtraForm(form_data)

            if form.is_valid():
                form.save()

        return redirect(weekoff)  

    else:
        form = ExtraForm()
        a = regmodel.objects.filter(status='Active')

        return render(request, 'weekoff.html', {'a': a, 'form': form})
    
# Weekoff display by admin    
from django.shortcuts import render
from .models import ExtraModel

def weekoffdisplayadmin(request):
    try:
        # Fetch all unique dates from the ExtraModel where status is 'Weekoff'
        unique_dates = ExtraModel.objects.filter(status='Weekoff').values_list('date', flat=True).distinct()

        # Create a list to store the data for each date
        weekoff_data_by_date = []

        # Iterate over unique dates
        for date in unique_dates:
            # Fetch records for the current date and status 'Weekoff'
            records_for_date = ExtraModel.objects.filter(date=date, status='Weekoff')

            # Extract the name and employeeid for each record and store in a list
            records_data = [{'name': record.name, 'employeeid': record.employeeid} for record in records_for_date]

            # Append the data for the current date to the list
            weekoff_data_by_date.append({'date': date, 'records': records_data})

        # Pass the data to the template for rendering
        return render(request, 'weekoffdisplayadmin.html', {'weekoff_data_by_date': weekoff_data_by_date})

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})
    
    
# Storing of sunday and public holiday into database by admin
from django.shortcuts import render, redirect
from .models import PublicholidaysModel , ExtraModel
from django.utils import timezone
from datetime import timedelta, date

def generate_sundays(year):
    # Get the first day of the year
    first_day = date(year, 1, 1)

    # Calculate the number of days to Sunday
    days_to_sunday = 6 - first_day.weekday()

    # Calculate the first Sunday of the year
    first_sunday = first_day + timedelta(days=days_to_sunday)

    # Generate all Sundays of the year
    sundays = [first_sunday + timedelta(weeks=i) for i in range((date(year, 12, 31) - first_sunday).days // 7 + 1)]

    return sundays

from django.shortcuts import render, redirect
from django.utils import timezone
from .models import PublicholidaysModel, ExtraModel, regmodel

def publicholidays(request):
    if request.method == 'POST':
        date_value = request.POST.get('date')
        status = request.POST.get('status')

        # Check if Sundays for the current year are already stored
        current_year = timezone.now().year
        existing_sundays_count = ExtraModel.objects.filter(date__year=current_year, status='Sunday').count()

        if existing_sundays_count == 0:
            # Generate and save Sundays for the current year
            sundays = generate_sundays(current_year)

            for sunday in sundays:
                # Save Sundays to ExtraModel
                employees = regmodel.objects.all()  
                for employee in employees:
                    sunday_data = ExtraModel(name=employee.name, employeeid=employee.employeeid, date=sunday, status='Sunday')
                    sunday_data.save()

        # Loop through each employee and save the data to both models
        employees = regmodel.objects.all() 
        for employee in employees:
            # Save public holidays to ExtraModel
            extra_data = ExtraModel(name=employee.name, employeeid=employee.employeeid, date=date_value, status='Holiday')
            extra_data.save()

            # Save public holidays to PublicholidaysModel
            public_holiday = PublicholidaysModel(date=date_value, status=status)
            public_holiday.save()

        return redirect(admindashboard)  

    # If it's not a POST request, check if Sundays for the current year are already stored
    current_year = timezone.now().year
    existing_sundays_count = ExtraModel.objects.filter(date__year=current_year, status='Sunday').count()

    if existing_sundays_count == 0:
        # If Sundays are not stored, generate and save them
        sundays = generate_sundays(current_year)

        for sunday in sundays:
            # Save Sundays to ExtraModel
            employees = regmodel.objects.all()  
            for employee in employees:
                sunday_data = ExtraModel(name=employee.name, employeeid=employee.employeeid, date=sunday, status='Sunday')
                sunday_data.save()

    return render(request, 'publicholidays.html')


# Public holidays display for admin
from django.shortcuts import render
from django.http import JsonResponse
from .models import PublicholidaysModel

from django.db.models import Count

def publicholidaysadmin(request):
    try:
        # Fetch all records from the PublicholidaysModel where day is not 'Sunday'
        # Group the records by date and count the occurrences of each date
        publicholidays = (
            PublicholidaysModel.objects.exclude(status='Sunday')
            .values('date','status')
            .annotate(date_count=Count('date'))
        )

        # Pass the grouped records to the template for rendering
        return render(request, 'publicholidaysadmin.html', {'publicholidays': publicholidays})

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})
    
# Productivity display for admin
from django.shortcuts import render
from .models import ProductivityModel

from django.shortcuts import render
from .models import ProductivityModel
from .models import regmodel

from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from .models import ProductivityModel

def display_productivity_data(request):
    # Retrieve all instances of ProductivityModel and reverse the queryset
    productivity_data = ProductivityModel.objects.all().order_by('-id')
    
    # Fetching names based on employee IDs
    employee_names = [(entry.employeeid, entry.name) for entry in regmodel.objects.all()]
    
    # Pagination
    paginator = Paginator(productivity_data, 10)  # 10 entries per page
    page = request.GET.get('page')
    try:
        productivity_data = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page.
        productivity_data = paginator.page(1)
    except EmptyPage:
        # If page is out of range (e.g. 9999), deliver last page of results.
        productivity_data = paginator.page(paginator.num_pages)
    
    # Pass the data and names to the template for rendering
    return render(request, 'productivityadmin.html', {'productivity_data': productivity_data, 'employee_names': employee_names})



def productivitymanager(request):
    current_user_id = request.session.get('id')
    if current_user_id:
        try:
            current_user = ManagerModel.objects.get(employeeid=current_user_id)
            managed_department = current_user.department
            
            # Fetch employees of the managed department
            managed_employees = regmodel.objects.filter(department=managed_department)
            
            # Retrieve productivity data for managed employees
            productivity_data = ProductivityModel.objects.filter(employeeid__in=managed_employees.values_list('employeeid', flat=True)).order_by('-id')
            
            # Fetching names based on employee IDs
            employee_names = regmodel.objects.filter(employeeid__in=managed_employees.values_list('employeeid', flat=True)).values_list('employeeid', 'name')
            
            # Pagination
            paginator = Paginator(productivity_data, 10)  # 10 entries per page
            page = request.GET.get('page')
            try:
                productivity_data = paginator.page(page)
            except PageNotAnInteger:
                # If page is not an integer, deliver first page.
                productivity_data = paginator.page(1)
            except EmptyPage:
                # If page is out of range (e.g. 9999), deliver last page of results.
                productivity_data = paginator.page(paginator.num_pages)
            
            # Pass the data and names to the template for rendering
            return render(request, 'productivitymanager.html', {'productivity_data': productivity_data, 'employee_names': employee_names})

        except ManagerModel.DoesNotExist:
            pass
    else:
        pass

    return render(request, 'productivitymanager.html', {'productivity_data': [], 'employee_names': []})

#Admin logout
from django.contrib.auth import logout

def adminlogout(request):
    logout(request)
    return redirect(adminlogin)

def managerlogout(request):
    logout(request)
    return redirect(managerlogin)

def terminated_employees(request):
    terminated_employees = regmodel.objects.filter(status='Terminated')
    return render(request, 'terminated.html', {'terminated_employees': terminated_employees})


from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.conf import settings
import os

def view_pdf(request, file_path):
    # Construct the full path to the PDF file
    pdf_path = os.path.join(settings.MEDIA_ROOT, file_path)

    # Check if the file exists
    if os.path.exists(pdf_path):
        # Open the PDF file in binary mode
        with open(pdf_path, 'rb') as pdf_file:
            # Set the content type for PDF
            response = HttpResponse(pdf_file.read(), content_type='application/pdf')
            # Set the HTTP headers for PDF file attachment
            response['Content-Disposition'] = 'inline; filename=' + os.path.basename(pdf_path)
            return response
    else:
        # Return 404 if the file does not exist
        return HttpResponse("File not found", status=404)
    
    
def change_password(request):
    if request.method == 'POST':
        id1 = request.session['id']
        employee = regmodel.objects.get(employeeid=id1)
        new_password = request.POST.get('newPassword')
        confirm_password = request.POST.get('confirmPassword')

        if new_password == confirm_password:
            # Update the password for the employee
            employee.password = new_password
            employee.save()
            return redirect(page3)  # Redirect to the profile page after password change
        else:
            # Handle password mismatch error
            error_message = "Passwords do not match. Please try again."
            return render(request, 'password_change_error.html', {'error_message': error_message})

    # If the request method is not POST, render the change password modal again
    return render(request, 'change_password_modal.html')


from django.shortcuts import render, redirect
from .models import ExtraModel, PublicholidaysModel

def delete_holiday(request, date):
    if request.method == 'POST':
        # Delete data from ExtraModel
        ExtraModel.objects.filter(date=date).delete()
        
        # Delete data from PublicholidaysModel
        PublicholidaysModel.objects.filter(date=date).delete()
        
        return redirect(publicholidaysadmin)
    
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth import authenticate

def forgotpassword(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        # Logic to send OTP to the email address
        messages.success(request, 'An OTP has been sent to your email address.')
        return redirect(forgotpassword)  # Redirect to the same page after sending OTP
    return render(request, 'forgotpassword.html')




def editemployeedetailsmanager(request, id):
    a = regmodel.objects.get(id=id)
    img = str(a.image).split('/')[-1]

    if request.method == 'POST':
        a.name = request.POST.get('name')
        a.employeeid = request.POST.get('employeeid')
        a.designation = request.POST.get('designation')
        a.email = request.POST.get('email')
        a.bloodgroup = request.POST.get('bloodgroup')
        a.phonenumber = request.POST.get('phonenumber')
        a.dateofbirth = request.POST.get('dateofbirth')
        a.joiningdate = request.POST.get('joiningdate')
        a.accountnumber = request.POST.get('accountnumber')
        a.bankname = request.POST.get('bankname')
        a.branch = request.POST.get('branch')
        a.ifsccode = request.POST.get('ifsccode')
        a.salary = request.POST.get('salary')
        a.logintime = request.POST.get('logintime')
        a.address = request.POST.get('address')
        a.department = request.POST.get('department')
        a.shifttime = request.POST.get('shifttime')
        a.maritalstatus = request.POST.get('maritalstatus')
        a.gender = request.POST.get('gender')

        # Handling file uploads
        if 'image' in request.FILES:
            if a.image:
                os.remove(a.image.path)
            a.image = request.FILES['image']

        # Update PDF files
        for field_name in ['idproof', 'educationalcertificate', 'workexperience', 'resume', 'others']:
            if field_name in request.FILES:
                file_field = request.FILES[field_name]
                if getattr(a, field_name):
                    os.remove(getattr(a, field_name).path)
                setattr(a, field_name, file_field)

        a.save()
        return redirect(employeedetailsman)
     # Fetching list of employees for reporting manager dropdown
    employees = regmodel.objects.filter(status__in=['Active', 'Inactive'])
    departments = DepartmentModel.objects.all()
    return render(request, 'editemployeedetailsmanager.html', {'a': a, 'img': img, 'employees': employees, 'departments':departments})



from django.shortcuts import render
from .models import DepartmentModel

from django.shortcuts import render

from .models import regmodel

from .models import DepartmentModel


# views.py

from django.shortcuts import render
from .models import DepartmentModel, regmodel  # Import the necessary models

from django.shortcuts import render
from .models import DepartmentModel, ManagerModel

from django.shortcuts import render
from .models import DepartmentModel, ManagerModel, regmodel
from django.http import JsonResponse

from django.shortcuts import render
from django.http import JsonResponse
from .models import DepartmentModel, ManagerModel, regmodel

def department_list(request):
    if request.method == 'POST':
        # Handle saving reporting manager
        employee_id = request.POST.get('employee_id')
        department_id = request.POST.get('department_id')
        manager = ManagerModel(employeeid=employee_id, department=department_id)
        manager.save()
        return JsonResponse({'message': 'Reporting manager saved successfully'})
    else:
        departments = DepartmentModel.objects.all()
        department_employees = {}

        # Fetch department managers
        for department in departments:
            manager = ManagerModel.objects.filter(department=department.id).first()
            department.manager = None
            if manager:
                employees = regmodel.objects.filter(employeeid=manager.employeeid)
                if employees.exists():
                    employee = employees.first()  # Choose one of the employees
                    department.manager = employee.name
                    department.manager_designation = employee.designation  # Fetch designation
            employees = regmodel.objects.filter(department=department.id)
            department_employees[department] = employees

        regmodel_list = regmodel.objects.all()  # Fetch all employees
        return render(request, 'department.html', {'department_employees': department_employees, 'regmodel_list': regmodel_list})



from django.shortcuts import redirect

def update_department(request):
    if request.method == 'POST':
        department_id = request.POST.get('department_id')
        new_department_name = request.POST.get('department_name')
        department = DepartmentModel.objects.get(id=department_id)
        department.department = new_department_name
        department.save()
        return redirect('department_list')  # Redirect to the department list page




from django.http import JsonResponse    
from .models import DepartmentModel, ManagerModel

def save_reporting_manager(request):
    if request.method == 'POST':
        employee_id = request.POST.get('employee_id')
        department_id = request.POST.get('department_id')
        
        # Check if a manager entry already exists for the department
        existing_manager = ManagerModel.objects.filter(department=department_id).first()
        
        if existing_manager:
            # Update the existing manager entry
            existing_manager.employeeid = employee_id
            existing_manager.save()
            return JsonResponse({'message': 'Reporting manager updated successfully'})
        else:
            # Create a new manager entry
            manager = ManagerModel(employeeid=employee_id, department=department_id)
            manager.save()
            return JsonResponse({'message': 'Reporting manager saved successfully'})
    else:
        return JsonResponse({'error': 'Invalid request method'})



# In views.py

from django.shortcuts import render
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .models import Asset

def assets(request):
    assets = Asset.objects.all()
    return render(request, 'assets.html', {'assets': assets})

@csrf_exempt
def save_assets(request):
    if request.method == "POST":
        asset_name = request.POST.get('assetName')
        asset_description = request.POST.get('assetDescription')

        # Create and save the asset to the database
        asset = Asset.objects.create(name=asset_name, description=asset_description)

        # Return a success response
        return JsonResponse({'success': True})
    else:
        return JsonResponse({'success': False, 'error': 'Invalid method'})

@csrf_exempt
def update_asset(request):
    if request.method == 'POST':
        asset_id = request.POST.get('id')
        asset_name = request.POST.get('name')
        asset_description = request.POST.get('description')
        
        # Update the asset in the database
        try:
            asset = Asset.objects.get(id=asset_id)
            asset.name = asset_name
            asset.description = asset_description
            asset.save()
            return JsonResponse({'message': 'Asset updated successfully'})
        except Asset.DoesNotExist:
            return JsonResponse({'error': 'Asset not found'})
    else:
        return JsonResponse({'error': 'Invalid request method'})

@csrf_exempt
def delete_asset(request):
    if request.method == 'POST':
        asset_id = request.POST.get('id')
        
        # Delete the asset from the database
        Asset.objects.filter(id=asset_id).delete()
        return JsonResponse({'message': 'Asset deleted successfully'})
    else:
        return JsonResponse({'error': 'Invalid request method'})


from django.shortcuts import render, redirect
from .models import CommentsModel

def save_comments(request):
    if request.method == 'POST':
        comments = request.POST.get('comments')
        employeeid = request.POST.get('employeeid')
        name = request.POST.get('name')
        # Assuming you have a model named CommentsModel to store comments
        CommentsModel.objects.create(employeeid=employeeid, name=name, comments=comments)
        return redirect(employeedetails)
    return redirect(employeedetails)  # Redirect to the same page if not a POST request


def get_comments(request, employeeid):
    comments = CommentsModel.objects.filter(employeeid=employeeid)
    comments_list = list(comments.values('name', 'comments'))
    return JsonResponse(comments_list, safe=False)



from django.shortcuts import redirect, get_object_or_404
from .models import ExcelModel, ExtraModel
from django.urls import reverse



def delete_entry(request, entry_id):
    if request.method == "POST":
        # Fetch the entry from ExcelModel
        entry = get_object_or_404(ExcelModel, id=entry_id)

        # Delete the corresponding data from ExtraModel for the same date and employee
        ExtraModel.objects.filter(employeeid=entry.employeeid, date=entry.date).delete()

        # Delete the entry from ExcelModel
        entry.delete()

    # Redirect to the page displaying the table
    return redirect(reverse(display_attendance_details))




















